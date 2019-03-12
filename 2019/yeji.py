#encoding utf-8
from copy import deepcopy
import xlrd, xlwt
import os
import numpy as np
import pandas as pd
import xlsxwriter
from xlutils.copy import copy 
from openpyxl import load_workbook


'''把各个店的数据读取到一个dataframe中，然后保存到一个CSV文件中，然后手动复制到区组统计中，
这样程序好写多了，操作也不麻烦，只要跑一次程序然后拷贝一下'''



rootdir='D:\\Users\\Administrator\\Desktop\\test'  # 放文件的目录
dest_file = '张磊区组业绩2019.03.10.xlsx'  # 需要统计的目标文件
sheet_name= '10'  # 表格名， 业绩中 表示日期的日




files = os.listdir(rootdir)
#shop_row_map = ['金创大厦','康桥万信酒店','川沙现代广场店','森兰商都','富海商务','高帆大厦','恒生万鹂','开文大厦','日月光','莲花科创园']


shop_row_map={
    '金创大厦': 0,     
    '日月光': 1,
    '开文大厦': 2,
    '莲花': 3,
    '富海': 4,
    '浦东软件园-1':5,
    '森兰商都': 6,
    '晓富金融': 7,
    '森兰国际': 8,
    '浦东软件园-2': 9,
    '高帆大厦': 10,
    '复旦科技园': 11,
    '如意智慧酒店': 12,
    '外高桥店': 13,
    '张江集电港': 14,
    '传奇广场': 15,
    '光大安石': 16,
    '海德堡': 17,
    '康桥万信': 18,
    '畅星大厦': 19,
    '中兴和泰': 20,
    '浦东机场': 21,
    '恒生万鹂': 22,
    '川沙现代广场': 23
}

df = pd.DataFrame(index=shop_row_map.keys(),columns=range(9))
#print(df)

dct = {}  # 把所有数据读入这个字典，然后把这个字典中的数据写到区组统计文件中
for filename in files:
    if filename == dest_file:  # 
        dest_file_path = os.path.join(rootdir, filename)
        continue
    # 遍历每个店的业绩文件，把数据放在一个大的字典里  格式   {店1：[一串数据], 店2：[一串数据],.......}
    # 然后遍历这个字典，把数据填到区组业绩中
    
    for short_name in shop_row_map.keys():
        #print(filename)
        if short_name in filename:
            
            filepath = os.path.join(rootdir, filename)

            #print(filepath)
            workbook = xlrd.open_workbook(filepath)

            tb = workbook.sheet_by_name(sheet_name)
            for i in range(1, 30):# 有些不一定写在第三行，所以要一行行检查，哪行有数据，记录哪行的数据
                one_data = [tb.cell(i,4).value, tb.cell(i,5).value, tb.cell(i,7).value,
                    tb.cell(i,9).value, tb.cell(i,10).value, tb.cell(i,11).value,
                    tb.cell(i,12).value, tb.cell(i,13).value, tb.cell(i,14).value]
                arr = np.array(one_data)
                #print(one_data, arr, arr.sum())
                lst = list(arr)
                #print(lst,type(lst[0]))
                if type(arr[0]) is np.float64: # 说明是数字，
                    #print(one_data,shop_row_map[short_name])
                    df.iloc[shop_row_map[short_name]] = one_data
                    #print(df)
                    #dct[short_name] = deepcopy(one_data)
                    break
                    
            #print(sheet.cell(3,2))  # cell(行，列)   从0 开始
            #print(sheet.col_values(2))
    

#print(dct)  # 虽然字典数据没错，但遍历重复了， 以后改



print(df)
df.to_csv('tmp.csv')
